import os
import logging
from datetime import datetime
from tempfile import gettempdir

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
import requests
import json

from flask_app.config.mysqlconnection import connectToMySQL

# Load .env for local dev
try:
    from dotenv import load_dotenv
    # Force .env values to override any existing env vars in shell
    load_dotenv(override=True)
except Exception:
    pass

# Uses Microsoft Graph directly for SharePoint uploads (no Office365 client)


logger = logging.getLogger("sharepoint_export")
if not logger.handlers:
    logging.basicConfig(level=logging.INFO)


DB_NAME = "man_hours"


class DatabaseExporter:
    """Encapsulates DB queries and Excel creation."""

    def __init__(self, db_name: str = DB_NAME):
        self.db_name = db_name

    def fetch_shifts_with_context(self, start_date: datetime, end_date: datetime):
        """Return shifts within date range, including user and job context."""
        query = (
            """
            SELECT 
                s.id AS shift_id,
                s.created_at AS shift_start,
                s.updated_at AS shift_end,
                TIMESTAMPDIFF(SECOND, s.created_at, COALESCE(s.updated_at, NOW())) AS duration_seconds,
                s.note,
                u.id AS user_id,
                u.first_name,
                u.last_name,
                u.email,
                u.department,
                j.id AS job_id,
                j.im_number,
                j.general_contractor,
                j.job_scope,
                j.estimated_hours,
                j.status AS job_status
            FROM shifts s
            JOIN users u ON s.user_id = u.id
            LEFT JOIN jobs j ON s.job_id = j.id
            WHERE s.created_at >= %s AND s.created_at <= %s
            ORDER BY s.created_at DESC
            """
        )
        params = (start_date, end_date)
        return connectToMySQL(self.db_name).query_db(query, params) or []

    def create_workbook(self, shifts_rows: list[dict]):
        """Create an Excel workbook with Shifts and Summary sheets."""
        wb = Workbook()

        # Shifts sheet
        ws = wb.active
        ws.title = "Shifts"
        if shifts_rows:
            df = pd.DataFrame(shifts_rows)
            for row in dataframe_to_rows(df, index=False, header=True):
                ws.append(row)
            self._format_table_sheet(ws)
        else:
            ws.append(["No shifts found for selected range"])

        # Summary sheet
        ws_sum = wb.create_sheet("Summary", 0)
        self._fill_summary(ws_sum, shifts_rows)

        return wb

    def _format_table_sheet(self, ws):
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
        # Best-effort auto width
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    def _fill_summary(self, ws, shifts_rows: list[dict]):
        total = len(shifts_rows)
        total_seconds = sum((r.get("duration_seconds") or 0) for r in shifts_rows)
        total_hours = round(total_seconds / 3600, 2)

        ws.append(["Island Time - Database Export Summary"])
        ws["A1"].font = Font(size=16, bold=True)
        ws.append([""])
        ws.append(["Exported At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws.append(["Shifts in Range", total])
        ws.append(["Total Shift Hours (approx)", total_hours])


def _upload_to_sharepoint(file_path: str) -> bool:
    """Upload a file to SharePoint using Microsoft Graph only (app-only auth)."""
    sp_url = os.getenv("SHAREPOINT_URL")
    sp_client_id = os.getenv("SHAREPOINT_CLIENT_ID")
    sp_client_secret = os.getenv("SHAREPOINT_CLIENT_SECRET")
    doc_lib = os.getenv("SHAREPOINT_DOC_LIBRARY", "Shared Documents")
    target_folder = os.getenv("SHAREPOINT_TARGET_FOLDER", "TimeTracker_Exports")

    if not (sp_url and sp_client_id and sp_client_secret):
        logger.error("Missing SHAREPOINT_URL, SHAREPOINT_CLIENT_ID, or SHAREPOINT_CLIENT_SECRET.")
        return False

    return _upload_via_graph_direct(file_path, sp_url, doc_lib, target_folder, sp_client_id, sp_client_secret)


def _upload_via_graph_direct(file_path: str, site_url: str, doc_library: str, target_folder: str, client_id: str, client_secret: str) -> bool:
    """Upload a file to SharePoint using Microsoft Graph with direct token request.

    - Uses the default Documents drive (sites/{site}/drive), or resolves a named library via /drives
    - Ensures the nested folder path exists
    - Uploads via simple PUT (works for files < 4MB)
    """
    try:
        from urllib.parse import urlparse

        parsed = urlparse(site_url)
        host = parsed.hostname
        site_path = parsed.path  # e.g., /sites/IslandMillworkOperations
        if not host or not site_path:
            logger.error("Invalid SHAREPOINT_URL; must include host and site path.")
            return False

        # Determine tenant authority for token issuance
        tenant_authority = os.getenv("AZURE_TENANT_ID") or f"{host.split('.')[0]}.onmicrosoft.com"
        token_url = f"https://login.microsoftonline.com/{tenant_authority}/oauth2/v2.0/token"
        token_data = {
            "grant_type": "client_credentials",
            "client_id": client_id,
            "client_secret": client_secret,
            "scope": "https://graph.microsoft.com/.default",
        }
        tok = requests.post(token_url, data=token_data)
        if tok.status_code != 200:
            logger.error(f"Failed to acquire Graph token: {tok.status_code} {tok.text}")
            return False
        access_token = tok.json().get("access_token")
        # Log token roles (app permissions) for diagnostics without exposing token content
        try:
            import base64, json as _json
            parts = access_token.split(".")
            if len(parts) >= 2:
                padded = parts[1] + "=" * (-len(parts[1]) % 4)
                payload = _json.loads(base64.urlsafe_b64decode(padded).decode("utf-8"))
                roles = payload.get("roles") or payload.get("scp")
                if roles:
                    logger.info(f"Graph token roles/scopes: {roles}")
        except Exception:
            pass
        headers_json = {"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"}

        # Resolve site ID
        site_api = f"https://graph.microsoft.com/v1.0/sites/{host}:{site_path}"
        sresp = requests.get(site_api, headers=headers_json)
        site_id = None
        if sresp.status_code == 200:
            site_id = sresp.json().get("id")
        else:
            logger.warning(f"Direct site resolve failed ({sresp.status_code}). Trying Graph search fallback...")
            # Fallback: use search to locate site by last segment
            last_seg = site_path.strip("/").split("/")[-1]
            search_url = f"https://graph.microsoft.com/v1.0/sites?search={last_seg}"
            s2 = requests.get(search_url, headers=headers_json)
            if s2.status_code == 200:
                candidates = s2.json().get("value", [])
                # Prefer exact path match when possible
                for c in candidates:
                    web_url = c.get("webUrl", "")
                    if web_url.endswith(site_path):
                        site_id = c.get("id")
                        break
                if not site_id and candidates:
                    # Fallback to first candidate
                    site_id = candidates[0].get("id")
                    logger.info("Using first site candidate from Graph search; verify if mismatch occurs.")
            if not site_id:
                logger.error(
                    f"Failed to resolve site via Graph. Direct: {sresp.status_code} {sresp.text} | "
                    f"Search: {s2.status_code if 's2' in locals() else 'n/a'} {s2.text if 's2' in locals() else ''}"
                )
                return False

        # Resolve drive (document library)
        drive_id = None
        if doc_library.strip().lower() == "shared documents":
            dresp = requests.get(f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive", headers=headers_json)
            if dresp.status_code == 200:
                drive_id = dresp.json().get("id")
            else:
                logger.error(f"Failed to get default drive: {dresp.status_code} {dresp.text}")
                return False
        else:
            # Enumerate drives and find by name
            dlist = requests.get(f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives", headers=headers_json)
            if dlist.status_code != 200:
                logger.error(f"Failed to list drives: {dlist.status_code} {dlist.text}")
                return False
            for drv in dlist.json().get("value", []):
                if drv.get("name") == doc_library:
                    drive_id = drv.get("id")
                    break
            if not drive_id:
                logger.error(f"Drive '{doc_library}' not found on site")
                return False

        # Ensure folder path exists
        folder_path = (target_folder or "").replace("\\", "/").strip().strip("/")
        current_path = ""
        if folder_path:
            for segment in folder_path.split("/"):
                parent_path = current_path
                current_path = f"{current_path}/{segment}" if current_path else segment
                # Check existence
                chk = requests.get(
                    f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{current_path}",
                    headers=headers_json,
                )
                if chk.status_code == 200:
                    continue
                # Create folder at appropriate parent
                children_url = (
                    f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root/children"
                    if not parent_path
                    else f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{parent_path}:/children"
                )
                create_body = {"name": segment, "folder": {}, "@microsoft.graph.conflictBehavior": "replace"}
                cre = requests.post(children_url, headers=headers_json, data=json.dumps(create_body))
                if cre.status_code not in (200, 201):
                    if cre.status_code == 409:
                        continue
                    logger.error(f"Failed to create folder '{segment}' under '{parent_path}': {cre.status_code} {cre.text}")
                    return False

        # Upload the file (simple PUT)
        with open(file_path, "rb") as fp:
            content = fp.read()
        file_name = os.path.basename(file_path)
        up_url = (
            f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{current_path}/{file_name}:/content"
            if current_path
            else f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{file_name}:/content"
        )
        up_headers = {"Authorization": f"Bearer {access_token}", "Content-Type": "application/octet-stream"}
        up = requests.put(up_url, headers=up_headers, data=content)
        if up.status_code not in (200, 201):
            logger.error(f"Graph upload failed: {up.status_code} {up.text}")
            return False
        logger.info("Graph upload succeeded.")
        return True
    except Exception as e:
        logger.exception(f"Graph upload error: {e}")
        return False


def export_database_to_sharepoint(start_date: datetime = None, end_date: datetime = None) -> str:
    """Export data in date range to Excel and upload to SharePoint."""
    try:
        # Default to July 1 – Aug 31, 2025 if no range provided
        if start_date is None:
            start_date = datetime(2025, 7, 1)
        if end_date is None:
            end_date = datetime(2025, 8, 31, 23, 59, 59)

        exporter = DatabaseExporter()
        logger.info(f"Fetching shifts from {start_date} to {end_date}...")
        shifts = exporter.fetch_shifts_with_context(start_date, end_date)

        logger.info("Creating Excel workbook...")
        wb = exporter.create_workbook(shifts)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"IslandTime_Export_{ts}.xlsx"
        local_dir = gettempdir()
        local_path = os.path.join(local_dir, filename)
        wb.save(local_path)
        logger.info(f"Saved export to {local_path}")

        uploaded = _upload_to_sharepoint(local_path)

        # Cleanup
        try:
            os.remove(local_path)
        except Exception:
            pass

        return "Export completed successfully" if uploaded else "Export failed during SharePoint upload"
    except Exception as e:
        logger.exception(f"Export process failed: {e}")
        return f"Error: {e}"


if __name__ == "__main__":
    print(export_database_to_sharepoint())
